
import os
import pandas as pd
from openpyxl import load_workbook

def log_metrics_to_excel(metrics, file="metrics_log.xlsx", sheet_name="Sheet1"):
    df = pd.DataFrame([metrics])

    if os.path.exists(file):
        # Get last row of the sheet to append data correctly
        book = load_workbook(file)
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            startrow = sheet.max_row
        else:
            startrow = 0

        with pd.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Write data starting after the last row, skip header
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
    else:
        # File doesn't exist — create with headers
        df.to_excel(file, index=False, sheet_name=sheet_name)
